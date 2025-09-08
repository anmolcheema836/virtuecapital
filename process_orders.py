import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import requests
import io

# --- CONFIGURATION: UPDATE THESE VALUES ---
# Based on your inspection of Orders.html, change these values.
# Example: If your order container is <div class="order-main">, set ORDER_CONTAINER_TAG = 'div' and ORDER_CONTAINER_CLASS = 'order-main'
ORDER_CONTAINER_TAG = 'div'      # The HTML tag for a single order container (e.g., 'div', 'tr')
ORDER_CONTAINER_CLASS = 'order' # The CSS class for that container

IMAGE_SELECTOR = 'img'           # The CSS selector to find the image within an order container
AMOUNT_SELECTOR = 'span.amount'  # The CSS selector to find the order amount

def process_orders_to_excel(html_file="Orders.html", excel_file="orders_output.xlsx"):
    """
    Parses an HTML file from AliExpress to extract order information and saves it into an Excel file.
    """
    if not os.path.exists(html_file):
        print(f"Error: The file '{html_file}' was not found in this directory.")
        return

    # Create a directory to store downloaded images if needed
    if not os.path.exists("product_images"):
        os.makedirs("product_images")

    # Read and parse the local HTML file
    with open(html_file, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "AliExpress Orders"

    # Find all order containers
    orders = soup.find_all(ORDER_CONTAINER_TAG, class_=ORDER_CONTAINER_CLASS)

    if not orders:
        print("Warning: No orders found. Please check your ORDER_CONTAINER_TAG and ORDER_CONTAINER_CLASS settings.")
        print(f"Searching for: <{ORDER_CONTAINER_TAG} class='{ORDER_CONTAINER_CLASS}'>")
        return

    print(f"Found {len(orders)} orders. Processing...")

    # Start writing data from the first row
    current_row = 1

    for index, order in enumerate(orders):
        # --- 1. Extract Order Amount ---
        amount_element = order.select_one(AMOUNT_SELECTOR)
        order_amount = amount_element.get_text(strip=True) if amount_element else "Not Found"

        # --- 2. Extract and Process Image ---
        img_element = order.select_one(IMAGE_SELECTOR)
        if img_element and img_element.has_attr('src'):
            img_url = img_element['src']
            
            try:
                # Handle web URLs
                if img_url.startswith('http'):
                    response = requests.get(img_url, stream=True)
                    response.raise_for_status()
                    img_data = response.content
                    image = Image(io.BytesIO(img_data))

                # Handle Base64 encoded images
                elif img_url.startswith('data:image'):
                    from base64 import b64decode
                    header, encoded = img_url.split(',', 1)
                    img_data = b64decode(encoded)
                    image = Image(io.BytesIO(img_data))
                
                # Insert the image into the Excel cell
                image.height = 75
                image.width = 75
                ws.add_image(image, f'A{current_row}')

            except Exception as e:
                ws[f'A{current_row}'] = f"Image Error: {e}"
                print(f"Could not process image for order {index + 1}: {e}")
        else:
            ws[f'A{current_row}'] = "No Image"

        # --- 3. Write Data to Excel ---
        ws[f'B{current_row}'] = order_amount
        ws.row_dimensions[current_row].height = 60 # Set row height for the image

        print(f"Processed order {index + 1}/{len(orders)}")
        current_row += 1

    # Adjust column widths for better presentation
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20

    # Save the Excel file
    try:
        wb.save(excel_file)
        print(f"\nSuccess! Your orders have been saved to '{excel_file}'")
    except Exception as e:
        print(f"\nError: Could not save the Excel file. Reason: {e}")


if __name__ == '__main__':
    # This is the main function that runs when you execute the script.
    process_orders_to_excel()